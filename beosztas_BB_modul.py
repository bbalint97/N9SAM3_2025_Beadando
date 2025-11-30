import subprocess
import sys

def bb_csomagellenorzes(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

bb_csomagellenorzes("pandas")
bb_csomagellenorzes("openpyxl")
bb_csomagellenorzes("tkinter")
bb_csomagellenorzes("datetime")

import pandas as pd
from openpyxl import load_workbook
import datetime
import calendar
import tkinter as tk
from tkinter import ttk


def bb_add_esemeny(ical, start_dt, end_dt, title):
    ical.extend([
        "BEGIN:VEVENT",
        f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}",
        f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}",
        f"SUMMARY:{title}",
        "END:VEVENT"
    ])


def bb_ics_generalas(eredmeny, ev, munkalap, output_file):
    ical = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//BeosztasGeneratorBB//HU"
    ]

    honap_szam = int(munkalap[-2:])
    _, napok_szama = calendar.monthrange(ev, honap_szam)

    for nap in range(1, napok_szama + 1):
        adat = eredmeny.get(nap)
        if not adat:
            continue

        ertek = adat["ertek"]
        szin = adat["szin"]

        if ertek not in ["", None]:
            if ertek == "8":
                start = datetime.datetime(ev, honap_szam, nap, 8, 0)
                end = datetime.datetime(ev, honap_szam, nap, 16, 0)
                bb_add_esemeny(ical, start, end, "Délelőtt (8-16)")

            elif ertek == "N":
                start = datetime.datetime(ev, honap_szam, nap, 7, 0)
                end = datetime.datetime(ev, honap_szam, nap, 19, 0)
                bb_add_esemeny(ical, start, end, "Nappal (7-19)")

            elif ertek == "É":
                start = datetime.datetime(ev, honap_szam, nap, 19, 0)
                end = datetime.datetime(ev, honap_szam, nap) + datetime.timedelta(days=1, hours=7)
                bb_add_esemeny(ical, start, end, "Éjszaka (19-07)")

            elif ertek in ["SZ8", "SZ12"]:
                start = datetime.datetime(ev, honap_szam, nap, 0, 0)
                end = datetime.datetime(ev, honap_szam, nap, 23, 59)
                bb_add_esemeny(ical, start, end, "Szabadnap")

        if szin == "FFFF0000":
            start = datetime.datetime(ev, honap_szam, nap, 19, 0)
            end = datetime.datetime(ev, honap_szam, nap) + datetime.timedelta(days=1, hours=7)
            bb_add_esemeny(ical, start, end, "Készenlét")

    ical.append("END:VCALENDAR")

    with open(output_file, "w", encoding="utf-8") as f:
        f.write("\n".join(ical))


class BB_AppGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("app")
        self.root.geometry("500x470")

        tk.Label(root, text="Válassz munkalapot (2025_XX):", anchor="w").pack(fill="x", pady=5)
        self.honap = tk.StringVar()

        try:
            wb = load_workbook("beosztas.xlsx", data_only=True)
            lapok = [name for name in wb.sheetnames if name.startswith("2025_")]
        except:
            lapok = []

        frame1 = tk.Frame(root)
        frame1.pack(fill="x", padx=10)

        self.honap_box = ttk.Combobox(frame1, textvariable=self.honap, values=lapok, justify="left")
        self.honap_box.pack(fill="x")
        self.honap_box.bind("<<ComboboxSelected>>", self.frissit_dolgozok)

        tk.Label(root, text="Válassz dolgozót:", anchor="w").pack(fill="x", pady=5)
        self.dolgozo_valtozo = tk.StringVar()

        frame2 = tk.Frame(root)
        frame2.pack(fill="x", padx=10)

        self.dolgozo_box = ttk.Combobox(frame2, textvariable=self.dolgozo_valtozo,
                                        justify="left", state="disabled")
        self.dolgozo_box.pack(fill="x")

        tk.Button(root, text="Futtatás", command=self.run).pack(pady=10)

        tk.Label(root, text="Státusz:", anchor="w").pack(fill="x")
        self.status = tk.Text(root, height=8)
        self.status.pack(fill="both", expand=True, padx=10, pady=5)

        tk.Label(root, text="Berczi Bálint – N9SAM3 – 2025",
                 fg="#777777", anchor="center").pack(fill="x", pady=5)


    def frissit_dolgozok(self, event):
        munkalap = self.honap.get()

        try:
            df = pd.read_excel("beosztas.xlsx", sheet_name=munkalap)
        except:
            return self.log("Hiba: A munkalap nem olvasható!")

        lista = []

        for idx, row in df.iterrows():
            nev = row["Név"]
            dsz = row["Dsz."]

            if pd.isna(nev) or pd.isna(dsz):
                continue

            dsz = int(float(dsz))
            lista.append(f"{nev} ({dsz})")

        self.dolgozo_box["values"] = lista
        self.dolgozo_box.config(state="readonly")
        self.log("Dolgozólista frissítve.")


    def run(self):
        munkalap = self.honap.get()
        kijelolt = self.dolgozo_valtozo.get()

        if munkalap == "":
            return self.log("Hiba: Nincs munkalap kiválasztva!")

        if kijelolt == "":
            return self.log("Hiba: Nincs dolgozó kiválasztva!")

        try:
            dsz = int(kijelolt.split("(")[1].split(")")[0])
        except:
            return self.log("Hiba a dolgozó formátumában!")

        try:
            df = pd.read_excel("beosztas.xlsx", sheet_name=munkalap)
        except:
            return self.log("Hiba: Munkalap beolvasása sikertelen!")

        sor = df.loc[df["Dsz."] == dsz]
        if sor.empty:
            return self.log("Hiba: Nincs ilyen dolgozó a munkalapon!")

        row_index = sor.index[0] + 2

        wb = load_workbook("beosztas.xlsx", data_only=True)
        ws = wb[munkalap]

        eredmeny = {}

        honap_szam = int(munkalap[-2:])
        _, napok_szama = calendar.monthrange(2025, honap_szam)

        for col_idx, nap in enumerate(range(1, napok_szama + 1), start=3):
            cell = ws.cell(row=row_index, column=col_idx)
            ertek = cell.value
            szin = cell.fill.fgColor.rgb
            eredmeny[nap] = {
                "ertek": "" if ertek is None else str(ertek),
                "szin": szin
            }

        output = f"beosztas_{dsz}_{honap_szam}.ics"

        try:
            bb_ics_generalas(eredmeny, 2025, munkalap, output)
        except Exception as e:
            return self.log(f"Hiba ICS generálás közben: {e}")

        self.log(f"Siker!\nICS fájl elkészült:\n{output}")


    def log(self, text):
        self.status.delete("1.0", tk.END)
        self.status.insert(tk.END, text)


def start_app():
    root = tk.Tk()
    BB_AppGUI(root)
    root.mainloop()
